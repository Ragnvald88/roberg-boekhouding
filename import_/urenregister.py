"""Eenmalige import van Urenregister.xlsm naar werkdagen tabel."""

import openpyxl
from pathlib import Path
from datetime import datetime
from database import get_db, get_klanten, add_klant

# Optional mapping from Excel klant names to cleaner names.
# Pass custom mapping via klant_mapping parameter, or leave empty to use Excel names as-is.
DEFAULT_KLANT_MAPPING: dict[str, str] = {}

# Codes die NIET meetellen voor urencriterium
ACHTERWACHT_CODES = {
    'ANW_DR_WEEKEND_ACHTERWACHT',
    'ANW_DR_WERKDAG_AVOND_ACHTERWACHT',
    'ANW_DR_WERKDAG_NACHT_ACHTERWACHT',
    'AW-WK-A', 'AW-WK-E', 'AW-WKND-A', 'AW1',
}


async def import_urenregister(xlsm_path: Path, db_path: Path,
                              klant_mapping: dict[str, str] = None) -> dict:
    """Import Urenregister.xlsm into werkdagen table.

    Returns dict with counts: imported, skipped, errors.
    """
    mapping = klant_mapping if klant_mapping is not None else DEFAULT_KLANT_MAPPING

    wb = openpyxl.load_workbook(str(xlsm_path), keep_vba=True, data_only=True)
    ws = wb['Urentabel']

    # Get existing klanten and build name→id mapping
    klanten = await get_klanten(db_path)
    klant_name_to_id = {k.naam: k.id for k in klanten}

    conn = await get_db(db_path)
    try:
        imported = 0
        skipped = 0
        errors = []

        for row_idx in range(2, ws.max_row + 1):
            datum_cell = ws.cell(row=row_idx, column=6).value  # F: Datum
            if datum_cell is None:
                continue

            # Parse datum
            if isinstance(datum_cell, datetime):
                datum = datum_cell.strftime('%Y-%m-%d')
            elif isinstance(datum_cell, str):
                try:
                    datum = datetime.strptime(datum_cell, '%Y-%m-%d').strftime('%Y-%m-%d')
                except ValueError:
                    datum = datum_cell
            else:
                skipped += 1
                continue

            # Get klant
            klant_excel = str(ws.cell(row=row_idx, column=8).value or '').strip()  # H
            if not klant_excel:
                skipped += 1
                continue

            klant_naam = mapping.get(klant_excel, klant_excel)

            # Create klant if not exists
            if klant_naam not in klant_name_to_id:
                tarief = ws.cell(row=row_idx, column=14).value or 0  # N: Uurtarief
                km = ws.cell(row=row_idx, column=13).value or 0  # M: Retourafstand
                klant_id_new = await _add_klant_via_conn(
                    conn, klant_naam, float(tarief), float(km)
                )
                klant_name_to_id[klant_naam] = klant_id_new

            klant_id = klant_name_to_id[klant_naam]

            # Parse fields
            code = str(ws.cell(row=row_idx, column=7).value or '').strip()  # G: CODE
            activiteit = str(ws.cell(row=row_idx, column=9).value or '').strip()  # I
            locatie = str(ws.cell(row=row_idx, column=10).value or '').strip()  # J
            uren = ws.cell(row=row_idx, column=11).value  # K: Uren
            if not uren or float(uren) <= 0:
                skipped += 1
                continue
            uren = float(uren)

            km = float(ws.cell(row=row_idx, column=13).value or 0)  # M: Retourafstand
            tarief = float(ws.cell(row=row_idx, column=14).value or 0)  # N: Uurtarief
            km_tarief = float(ws.cell(row=row_idx, column=15).value or 0.23)  # O
            status = str(ws.cell(row=row_idx, column=19).value or '').strip().lower()  # S
            factuurnummer = str(ws.cell(row=row_idx, column=20).value or '').strip()  # T
            opmerking = str(ws.cell(row=row_idx, column=21).value or '').strip()  # U

            # Map status
            if status == 'betaald':
                db_status = 'betaald'
            elif status == 'gefactureerd':
                db_status = 'gefactureerd'
            elif status == 'niet van toepassing':
                db_status = 'ongefactureerd'
            else:
                db_status = 'ongefactureerd'

            # Urennorm: achterwacht codes don't count
            urennorm = 0 if code in ACHTERWACHT_CODES else 1

            # Also don't count if code contains ACHTERWACHT
            if 'ACHTERWACHT' in code.upper():
                urennorm = 0

            try:
                await conn.execute(
                    """INSERT INTO werkdagen
                       (datum, klant_id, code, activiteit, locatie, uren, km,
                        tarief, km_tarief, status, factuurnummer, opmerking, urennorm)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (datum, klant_id, code, activiteit, locatie, uren, km,
                     tarief, km_tarief, db_status, factuurnummer, opmerking, urennorm)
                )
                imported += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {e}")

        await conn.commit()
        return {'imported': imported, 'skipped': skipped, 'errors': errors}
    finally:
        await conn.close()


async def _add_klant_via_conn(conn, naam: str, tarief: float, km: float) -> int:
    """Add a klant directly via connection (during import)."""
    cursor = await conn.execute(
        "INSERT INTO klanten (naam, tarief_uur, retour_km, adres, actief) "
        "VALUES (?, ?, ?, '', 0)",
        (naam, tarief, km)
    )
    await conn.commit()
    return cursor.lastrowid


async def import_facturen_from_pdfs(inkomsten_base: Path, db_path: Path) -> dict:
    """Scan year/Inkomsten directories for PDF invoices, create factuur records.

    Expected filename patterns:
    - YYYY-NNN_Klant.pdf
    - YYYY-NNN Klant.pdf
    """
    import re

    klanten = await get_klanten(db_path)
    klant_name_to_id = {k.naam: k.id for k in klanten}

    conn = await get_db(db_path)
    try:
        imported = 0
        skipped = 0

        for year_dir in sorted(inkomsten_base.iterdir()):
            if not year_dir.is_dir() or not year_dir.name.isdigit():
                continue

            inkomsten_dir = year_dir / "Inkomsten"
            if not inkomsten_dir.exists():
                continue

            for pdf in sorted(inkomsten_dir.rglob("*.pdf")):
                # Parse YYYY-NNN from filename (supports 2-5+ digit numbers)
                # Formats: 2024-001_Klant.pdf, 2023-00001.pdf, 2023-016.pdf, 2023-09_DDG.pdf
                match = re.match(r'(\d{4}-\d{2,5})(?:[_\s](.+))?\.pdf', pdf.name)
                if not match:
                    skipped += 1
                    continue

                raw_nummer = match.group(1)
                # Normalize to YYYY-NNN (3 digits)
                year_part, num_part = raw_nummer.split('-')
                nummer = f"{year_part}-{int(num_part):03d}"
                klant_hint = (match.group(2) or '').replace('_', ' ').strip()

                # Check if factuur already exists
                cursor = await conn.execute(
                    "SELECT id FROM facturen WHERE nummer = ?", (nummer,)
                )
                if await cursor.fetchone():
                    skipped += 1
                    continue

                # Try to match klant
                klant_id = None
                for kn, kid in klant_name_to_id.items():
                    if klant_hint.lower() in kn.lower() or kn.lower() in klant_hint.lower():
                        klant_id = kid
                        break

                if not klant_id:
                    # Use first klant as fallback
                    klant_id = next(iter(klant_name_to_id.values()), 1)

                jaar = int(nummer[:4])
                await conn.execute(
                    """INSERT INTO facturen
                       (nummer, klant_id, datum, totaal_bedrag, pdf_pad, betaald, type)
                       VALUES (?, ?, ?, 0, ?, 1, 'factuur')""",
                    (nummer, klant_id, f"{jaar}-01-01", str(pdf))
                )
                imported += 1

        await conn.commit()
        return {'imported': imported, 'skipped': skipped}
    finally:
        await conn.close()
